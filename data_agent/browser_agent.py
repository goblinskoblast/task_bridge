from __future__ import annotations

import asyncio
import base64
import csv
import hashlib
import io
import json
import logging
import os
import re
import tempfile
from pathlib import Path
from typing import Any, Awaitable, Callable, Optional

import openpyxl
import xlrd
from PIL import Image

from bot.ai_provider import get_ai_provider
from email_integration.encryption import decrypt_password

logger = logging.getLogger(__name__)

ProgressCallback = Callable[[str], Awaitable[None]]


class BrowserAgent:
    MAX_STEPS = 15
    TIMEOUT_TOTAL = 120
    TIMEOUT_PAGE = 8000
    TIMEOUT_CLICK = 3000
    VIEWPORT_W = 1440
    VIEWPORT_H = 900
    SCREENSHOT_W = 1280
    SCREENSHOT_H = 800
    STUCK_THRESHOLD = 3

    NAVIGATOR_SYSTEM_PROMPT = """Ты Browser Agent для TaskBridge.

Отвечай строго JSON одним из вариантов:
1. {"action": "done", "data": "..."}
2. {"action": "click", "x": 640, "y": 380, "selector": "button.submit", "reason": "..."}
3. {"action": "fill", "x": 320, "y": 210, "selector": "input[name=login]", "value": "text", "reason": "..."}
4. {"action": "navigate", "url": "https://...", "reason": "..."}
5. {"action": "wait", "reason": "..."}

Приоритеты:
- Сначала обработай popup, cookie banner, предупреждение, форму подтверждения.
- Если видишь 2FA/SMS/код подтверждения, верни {"action":"done","data":"ТРЕБУЕТСЯ_2FA"}.
- Если видишь страницу ошибки доступа, верни {"action":"done","data":"ОШИБКА_ДОСТУПА: ..."}.
- Для кликов всегда возвращай и x/y, и selector.
- Если видишь кнопку экспорт/выгрузить/excel/download, предпочитай нажать её.
- Если данные уже на экране, верни action=done и кратко опиши найденное.
"""

    def __init__(self) -> None:
        self._last_download: Optional[str] = None

    def _describe_action(self, action: dict | None) -> str:
        if not action:
            return "none"
        action_name = str(action.get("action") or "unknown")
        selector = str(action.get("selector") or "").strip()
        reason = str(action.get("reason") or "").strip()
        parts = [action_name]
        if selector:
            parts.append(f"selector={selector}")
        if reason:
            parts.append(f"reason={reason[:120]}")
        return " | ".join(parts)

    def _build_runtime_diagnostic(
        self,
        *,
        stage: str,
        url: str,
        detail: str,
        last_action: str = "",
    ) -> str:
        message = f"BROWSER_AGENT_DIAGNOSTIC stage={stage} url={url or '-'} detail={detail}"
        if last_action:
            message += f" last_action={last_action}"
        return message[:1200]

    async def extract_data(
        self,
        url: str,
        username: str | None,
        encrypted_password: str | None,
        user_task: str,
        progress_callback: Optional[ProgressCallback] = None,
    ) -> str:
        password = decrypt_password(encrypted_password) if encrypted_password else ""
        username = username or ""
        logger.info(
            "BrowserAgent extract_data start url=%s has_credentials=%s task=%s",
            url,
            bool(username or password),
            user_task[:300],
        )
        try:
            from playwright.async_api import async_playwright
        except ImportError as exc:
            raise RuntimeError("Playwright is not installed") from exc

        async def _run() -> str:
            async with async_playwright() as playwright:
                browser = await playwright.chromium.launch(
                    headless=True,
                    args=[
                        "--disable-blink-features=AutomationControlled",
                        "--disable-dev-shm-usage",
                        "--no-sandbox",
                    ],
                )
                context = await browser.new_context(
                    viewport={"width": self.VIEWPORT_W, "height": self.VIEWPORT_H},
                    user_agent=(
                        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                        "AppleWebKit/537.36 (KHTML, like Gecko) "
                        "Chrome/120.0.0.0 Safari/537.36"
                    ),
                    locale="ru-RU",
                    timezone_id="Europe/Moscow",
                    accept_downloads=True,
                )
                await context.add_init_script(
                    """
                    Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
                    Object.defineProperty(navigator, 'platform', { get: () => 'Win32' });
                    Object.defineProperty(navigator, 'language', { get: () => 'ru-RU' });
                    Object.defineProperty(navigator, 'languages', { get: () => ['ru-RU', 'ru', 'en-US', 'en'] });
                    Object.defineProperty(navigator, 'plugins', { get: () => [1, 2, 3, 4, 5] });
                    window.chrome = window.chrome || { runtime: {} };
                    """
                )
                await context.set_extra_http_headers(
                    {
                        "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
                        "Upgrade-Insecure-Requests": "1",
                        "DNT": "1",
                    }
                )
                page = await context.new_page()
                stage = "page_created"

                try:
                    if progress_callback:
                        await progress_callback("Подключаюсь к внешней системе...")

                    stage = "goto"
                    await page.goto(url, wait_until="domcontentloaded", timeout=self.TIMEOUT_PAGE)
                    await self._safe_wait(page)
                    logger.info("BrowserAgent page loaded url=%s", page.url)

                    stage = "login"
                    login_result = await self._login(page, username, password, progress_callback)
                    logger.info("BrowserAgent login_result=%s", login_result)
                    if not login_result.get("success"):
                        return self._build_runtime_diagnostic(
                            stage=stage,
                            url=page.url,
                            detail=login_result.get("error", "Не удалось выполнить вход в систему."),
                        )

                    stage = "navigate_extract"
                    return await self._navigate_and_extract(page, user_task, progress_callback)
                except Exception as exc:
                    diagnostic = self._build_runtime_diagnostic(
                        stage=stage,
                        url=getattr(page, "url", url),
                        detail=str(exc),
                    )
                    logger.error("BrowserAgent runtime failure %s", diagnostic, exc_info=True)
                    raise RuntimeError(diagnostic) from exc
                finally:
                    await context.close()
                    await browser.close()

        return await asyncio.wait_for(_run(), timeout=self.TIMEOUT_TOTAL)

    async def _login(
        self,
        page: Any,
        username: str,
        password: str,
        progress_callback: Optional[ProgressCallback] = None,
    ) -> dict:
        if not username and not password:
            logger.info("BrowserAgent login skipped: no credentials")
            return {"success": True, "mode": "skip_login_no_credentials"}

        username_selectors = [
            "input[name='username']",
            "input[name='login']",
            "input[type='email']",
            "input[id*='login']",
            "input[id*='user']",
            "input[placeholder*='логин']",
            "input[placeholder*='email']",
        ]
        password_selectors = [
            "input[type='password']",
            "input[name='password']",
            "input[id*='password']",
            "input[placeholder*='пароль']",
        ]
        submit_selectors = [
            "button[type='submit']",
            "input[type='submit']",
            "button:has-text('Войти')",
            "button:has-text('Login')",
            "button:has-text('Sign in')",
        ]

        username_selector = await self._first_existing_selector(page, username_selectors)
        password_selector = await self._first_existing_selector(page, password_selectors)

        if not username_selector or not password_selector:
            logger.info("BrowserAgent login form not found, proceeding without form login")
            return {"success": True, "mode": "skip_login_form_not_found"}

        if progress_callback:
            await progress_callback("Выполняю вход...")

        await page.fill(username_selector, username)
        await page.fill(password_selector, password)

        submit_selector = await self._first_existing_selector(page, submit_selectors)
        if submit_selector:
            await page.click(submit_selector, timeout=self.TIMEOUT_CLICK)
        else:
            await page.keyboard.press("Enter")

        await self._safe_wait(page)

        special_state = await self._detect_special_state(page)
        if special_state:
            return {"success": False, "error": special_state}

        return {"success": True, "mode": "form_login"}

    async def _navigate_and_extract(
        self,
        page: Any,
        user_task: str,
        progress_callback: Optional[ProgressCallback] = None,
    ) -> str:
        screenshot_hashes: list[str] = []
        last_action_summary = ""
        self._last_download = None
        page.on("download", lambda download: asyncio.create_task(self._save_download(download)))

        for step in range(self.MAX_STEPS):
            await self._safe_wait(page)
            logger.info("BrowserAgent step=%s current_url=%s", step + 1, page.url)

            dismissed_overlay = await self._dismiss_common_overlays(page)
            if dismissed_overlay:
                logger.info("BrowserAgent dismissed common overlay on step=%s", step + 1)
                await self._safe_wait(page)

            if self._last_download:
                if progress_callback:
                    await progress_callback("Файл скачан, читаю данные...")
                file_path = self._last_download
                self._last_download = None
                return await self._parse_downloaded_file(file_path)

            special_state = await self._detect_special_state(page)
            if special_state:
                return special_state

            screenshot_b64 = await self._screenshot_b64(page)
            current_hash = hashlib.md5(screenshot_b64.encode()).hexdigest()
            screenshot_hashes.append(current_hash)
            if len(screenshot_hashes) >= self.STUCK_THRESHOLD and len(set(screenshot_hashes[-self.STUCK_THRESHOLD:])) == 1:
                if progress_callback:
                    await progress_callback("Страница не меняется, пробую выйти из зависшего состояния...")
                await page.keyboard.press("Escape")
                await self._safe_wait(page)
                screenshot_hashes.clear()

            if progress_callback and step > 0:
                await progress_callback(f"Шаг {step}/{self.MAX_STEPS}...")

            action = await self._decide_next_action(page, screenshot_b64, user_task, step)
            last_action_summary = self._describe_action(action)
            logger.info("BrowserAgent action step=%s action=%s", step + 1, action)
            result = await self._execute_action(page, action)

            if result is not None:
                return result

        return self._build_runtime_diagnostic(
            stage="max_steps",
            url=page.url,
            detail=f"Не удалось собрать данные за {self.MAX_STEPS} шагов",
            last_action=last_action_summary,
        )

    async def _decide_next_action(self, page: Any, screenshot_b64: str, user_task: str, step: int) -> dict:
        visible_text = await page.locator("body").inner_text()
        visible_text = visible_text[:4000]
        current_url = page.url

        try:
            provider = get_ai_provider()
            result = await provider.analyze_message(
                messages=[
                    {"role": "system", "content": self.NAVIGATOR_SYSTEM_PROMPT},
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "text",
                                "text": (
                                    f"USER TASK:\n{user_task}\n\n"
                                    f"CURRENT URL:\n{current_url}\n\n"
                                    f"STEP: {step + 1}/{self.MAX_STEPS}\n\n"
                                    f"VISIBLE TEXT:\n{visible_text}\n\n"
                                    "Ответь только JSON."
                                ),
                            },
                            {
                                "type": "image_url",
                                "image_url": {
                                    "url": f"data:image/png;base64,{screenshot_b64}"
                                },
                            },
                        ],
                    },
                ],
                temperature=0.1,
                max_tokens=350,
                response_format={"type": "json_object"},
            )
            if isinstance(result, dict) and result.get("action"):
                logger.info("BrowserAgent AI action selected step=%s action=%s", step + 1, result)
                return result
        except Exception as exc:
            logger.warning("BrowserAgent AI action fallback used: %s", exc)

        return await self._fallback_action(page, user_task, visible_text)

    async def _fallback_action(self, page: Any, user_task: str, visible_text: str) -> dict:
        lowered_task = user_task.lower()
        page_issue = self._classify_page_issue_text(visible_text)
        if page_issue:
            return {"action": "done", "data": page_issue}

        dismiss_candidates = [
            "button:has-text('Принять')",
            "button:has-text('Accept')",
            "button:has-text('Ок')",
            "button:has-text('OK')",
            "button:has-text('Понятно')",
            "button:has-text('Закрыть')",
            "[role='button']:has-text('Принять')",
            "[role='button']:has-text('Accept')",
            "text=Принять",
            "text=Accept",
        ]
        dismiss_selector = await self._first_existing_selector(page, dismiss_candidates)
        if dismiss_selector:
            box = await self._selector_center(page, dismiss_selector)
            return {
                "action": "click",
                "selector": dismiss_selector,
                "x": box["x"] if box else self.VIEWPORT_W // 2,
                "y": box["y"] if box else self.VIEWPORT_H // 2,
                "reason": "Похоже на popup/cookie banner",
            }

        export_candidates = [
            "button:has-text('Экспорт')",
            "button:has-text('Выгрузить')",
            "button:has-text('Excel')",
            "a:has-text('Экспорт')",
            "a:has-text('Выгрузить')",
            "a:has-text('Excel')",
            "text=Export",
            "text=Download",
        ]
        if any(token in lowered_task for token in ["отчет", "отчёт", "report", "выручк", "excel", "скач", "выгруз"]):
            selector = await self._first_existing_selector(page, export_candidates)
            if selector:
                box = await self._selector_center(page, selector)
                return {
                    "action": "click",
                    "selector": selector,
                    "x": box["x"] if box else self.VIEWPORT_W // 2,
                    "y": box["y"] if box else self.VIEWPORT_H // 2,
                    "reason": "Найдена кнопка экспорта",
                }

        if len(visible_text.strip()) > 100 and not self._looks_like_auth_screen(visible_text):
            return {"action": "done", "data": visible_text[:3000]}

        return {"action": "wait", "reason": "Ожидаю стабилизации страницы"}

    async def _execute_action(self, page: Any, action: dict) -> Optional[str]:
        action_name = action.get("action")
        logger.info("BrowserAgent execute action=%s url=%s", action_name, page.url)

        if action_name == "done":
            return str(action.get("data", "")).strip() or "Browser Agent завершил работу без данных."

        if action_name == "navigate":
            await page.goto(action.get("url", page.url), wait_until="domcontentloaded", timeout=self.TIMEOUT_PAGE)
            await self._safe_wait(page)
            return None

        if action_name == "wait":
            await self._safe_wait(page)
            return None

        if action_name == "click":
            clicked = False
            if action.get("x") is not None and action.get("y") is not None:
                try:
                    await page.mouse.click(action["x"], action["y"])
                    clicked = True
                except Exception:
                    pass
            selector = action.get("selector")
            if not clicked and selector:
                try:
                    await page.click(selector, timeout=self.TIMEOUT_CLICK)
                    clicked = True
                except Exception:
                    pass
            if not clicked and selector:
                try:
                    await page.evaluate(
                        """(sel) => {
                            const el = document.querySelector(sel);
                            if (el) el.click();
                        }""",
                        selector,
                    )
                    clicked = True
                except Exception:
                    pass
            if clicked:
                await self._safe_wait(page)
            else:
                logger.warning("BrowserAgent click action had no effect action=%s url=%s", action, page.url)
            return None

        if action_name == "fill":
            filled = False
            selector = action.get("selector")
            value = str(action.get("value", ""))
            if selector:
                try:
                    await page.fill(selector, value, timeout=self.TIMEOUT_CLICK)
                    filled = True
                except Exception:
                    pass
            if not filled and action.get("x") is not None and action.get("y") is not None:
                try:
                    await page.mouse.click(action["x"], action["y"])
                    await page.keyboard.press("Control+A")
                    await page.keyboard.type(value)
                    filled = True
                except Exception:
                    pass
            if filled:
                await self._safe_wait(page)
            else:
                logger.warning("BrowserAgent fill action had no effect action=%s url=%s", action, page.url)
            return None

        return None

    async def _detect_special_state(self, page: Any) -> Optional[str]:
        lowered_url = str(getattr(page, "url", "") or "").lower()
        if "2gis.ru/xpvnsulc" in lowered_url:
            return "ОШИБКА_ДОСТУПА: 2GIS временно отклонил автоматический запрос"
        text = await page.locator("body").inner_text()
        return self._classify_page_issue_text(text)

    async def _safe_wait(self, page: Any, timeout: int = TIMEOUT_PAGE) -> None:
        try:
            await page.wait_for_load_state("domcontentloaded", timeout=timeout)
        except Exception:
            pass

        spinner_selectors = [
            ".loading", ".spinner", ".loader",
            "[class*='loading']", "[class*='spinner']",
            ".x-mask", ".v-progress-circular", ".ant-spin", ".el-loading-mask",
        ]
        for selector in spinner_selectors:
            try:
                await page.wait_for_selector(selector, state="hidden", timeout=3000)
            except Exception:
                pass

        await asyncio.sleep(0.8)

    async def _screenshot_b64(self, page: Any) -> str:
        screenshot = await page.screenshot(full_page=False)
        image = Image.open(io.BytesIO(screenshot))
        image = image.resize((self.SCREENSHOT_W, self.SCREENSHOT_H), Image.LANCZOS)
        buffer = io.BytesIO()
        image.save(buffer, format="PNG", optimize=True)
        return base64.b64encode(buffer.getvalue()).decode("utf-8")

    async def _first_existing_selector(self, page: Any, selectors: list[str]) -> Optional[str]:
        for selector in selectors:
            try:
                locator = page.locator(selector)
                count = min(await locator.count(), 5)
                for idx in range(count):
                    try:
                        if await locator.nth(idx).is_visible():
                            return selector
                    except Exception:
                        continue
                if await locator.count() > 0:
                    return selector
            except Exception:
                continue
        return None

    async def _selector_center(self, page: Any, selector: str) -> Optional[dict]:
        try:
            box = await page.locator(selector).first.bounding_box()
            if not box:
                return None
            return {
                "x": int(box["x"] + box["width"] / 2),
                "y": int(box["y"] + box["height"] / 2),
            }
        except Exception:
            return None

    async def _save_download(self, download: Any) -> None:
        suffix = Path(download.suggested_filename).suffix or ".xlsx"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp_path = tmp.name
        await download.save_as(tmp_path)
        self._last_download = tmp_path

    async def _parse_downloaded_file(self, file_path: str) -> str:
        try:
            if file_path.endswith((".csv", ".tsv", ".txt")):
                return self._parse_delimited_text(file_path)
            if file_path.endswith(".xls"):
                return self._parse_xls(file_path)
            return self._parse_xlsx(file_path)
        except Exception as exc:
            return f"Файл скачан, но не удалось прочитать: {exc}"
        finally:
            try:
                os.remove(file_path)
            except OSError:
                pass

    def _parse_xlsx(self, file_path: str) -> str:
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        lines: list[str] = []
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            lines.append(f"[Лист: {sheet_name}]")
            row_count = 0
            for row in worksheet.iter_rows(values_only=True):
                if row_count >= 200:
                    lines.append("... (обрезано, больше 200 строк)")
                    break
                if all(value is None for value in row):
                    continue
                lines.append("\t".join("" if value is None else str(value) for value in row))
                row_count += 1
        return "\n".join(lines)

    def _parse_xls(self, file_path: str) -> str:
        workbook = xlrd.open_workbook(file_path)
        lines: list[str] = []
        for sheet in workbook.sheets():
            lines.append(f"[Лист: {sheet.name}]")
            for index in range(min(sheet.nrows, 200)):
                row = sheet.row_values(index)
                if all(value in ("", None) for value in row):
                    continue
                lines.append("\t".join(str(value) for value in row))
        return "\n".join(lines)

    def _parse_delimited_text(self, file_path: str) -> str:
        with open(file_path, "r", encoding="utf-8-sig", newline="") as handle:
            sample = handle.read(2048)
            handle.seek(0)
            delimiter = ";"
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                delimiter = dialect.delimiter
            except Exception:
                if sample.count(",") > sample.count(";"):
                    delimiter = ","
                elif sample.count("\t") > 0:
                    delimiter = "\t"

            reader = csv.reader(handle, delimiter=delimiter)
            lines: list[str] = []
            for idx, row in enumerate(reader):
                if idx >= 200:
                    lines.append("... (обрезано, больше 200 строк)")
                    break
                if not any((cell or "").strip() for cell in row):
                    continue
                lines.append("\t".join((cell or "").strip() for cell in row))
            return "\n".join(lines)

    def _classify_page_issue_text(self, text: str) -> Optional[str]:
        lowered = re.sub(r"\s+", " ", (text or "").lower()).strip()
        if not lowered:
            return None
        if any(token in lowered for token in ["код подтверждения", "sms", "2fa", "двухфактор", "verification code", "one-time code"]):
            return "ТРЕБУЕТСЯ_2FA"
        if any(token in lowered for token in ["captcha", "капча", "я не робот", "i am not a robot"]):
            return "ТРЕБУЕТСЯ_КАПЧА"
        if any(token in lowered for token in ["неверный пароль", "неверный логин", "invalid credentials", "wrong password", "incorrect password", "логин или пароль"]):
            return "ОШИБКА_АВТОРИЗАЦИИ: не удалось пройти вход"
        if any(token in lowered for token in ["нет доступа", "access denied", "403", "forbidden", "permission denied"]):
            return "ОШИБКА_ДОСТУПА: обнаружена страница отказа в доступе"
        return None

    def _looks_like_auth_screen(self, text: str) -> bool:
        lowered = re.sub(r"\s+", " ", (text or "").lower()).strip()
        auth_markers = [
            "войти",
            "login",
            "sign in",
            "password",
            "пароль",
            "username",
            "логин",
            "email",
            "вход в систему",
            "авторизация",
        ]
        return sum(1 for marker in auth_markers if marker in lowered) >= 2

    async def _dismiss_common_overlays(self, page: Any) -> bool:
        selectors = [
            "button:has-text('Принять')",
            "button:has-text('Accept')",
            "button:has-text('Согласен')",
            "button:has-text('Понятно')",
            "button:has-text('Закрыть')",
            "[role='button']:has-text('Принять')",
            "[role='button']:has-text('Accept')",
        ]
        selector = await self._first_existing_selector(page, selectors)
        if not selector:
            return False
        try:
            await page.locator(selector).first.click(timeout=self.TIMEOUT_CLICK)
            return True
        except Exception:
            return False


browser_agent = BrowserAgent()
